#!/usr/bin/env python3
"""
Municipal Code Review Tool
Compares a city's municipal code against MTAS model code and a standard checklist.

Directory layout:
  mtas/title_1.txt ... title_20.txt   (MTAS model, set up once)
  cities/<city>/title_1.txt ...       (city code, one folder per city)
  output/                             (Excel results written here)

Usage:
  python review.py --city Memphis --titles 1,3,6
  python review.py --city Memphis            # all 20 titles
"""

import argparse
import json
import os
import sys
from pathlib import Path
from datetime import datetime
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Standard checklist baked in from Legal and In-House Review Tool.xlsx
# ---------------------------------------------------------------------------
CHECKLIST = {
  "1": [
    {"chapter_title": None, "section": "1-101", "section_title": "Time and place of regular meetings", "comment": "Please review this section to verify this is still the correct meeting date and time or specify any revisions we should make."},
    {"chapter_title": None, "section": "1-102", "section_title": "Order of business", "comment": "Recommend using the model code section to include the public comment period as required by TCA § 8-44-112. (If a unique order of business is in place, recommend the public comment period with footnote. Also, “Grievances” is often recommended to be replaced by “Public Comment.”)"},
    {"chapter_title": None, "section": "1-103", "section_title": "General rules of order", "comment": "Recommend revising to remove the RRONR edition number and use the following language- “Robert’s Rules of Order, Newly Revised, current edition”."},
    {"chapter_title": None, "section": "1-609", "section_title": "Outside Employment", "comment": "Recommend replacing with the model. “An official or employee may not accept or continue any outside employment if the work unreasonably inhibits the performance of any affirmative duty of the municipal position or conflicts with any provision of the municipality’s charter or any ordinance or policy.” Your policy can then be included with the personnel policy. I also Recommend the authorization in policy be granted by the department head, city manager, or mayor."}
  ],
  "2": [
    {"chapter_title": None, "section": "2-000", "section_title": None, "comment": "Beer Board should be in title 8, chapter 2, not here."},
    {"chapter_title": None, "section": "2-000", "section_title": None, "comment": "Board of Zoning Appeals should be in title 14, chapter 1, not here."}
  ],
  "3": [
    {"chapter_title": "Bonds and Appeals", "section": "3-000", "section_title": "entire chapter", "comment": "Recommend this chapter be replaced in full with the MTAS model Title 3 Chapter 4."},
    {"chapter_title": "Bonds and Appeals", "section": "3-400", "section_title": "entire chapter", "comment": "Recommend using the MTAS model chapter \"Bonds and Appeals\" for current compliance with current statutory provisions."},
    {"chapter_title": "Court Administration", "section": "3-000", "section_title": "Imposition of penalties and costs.", "comment": "Please provide the information to complete the blanks in this section."},
    {"chapter_title": "Court Administration", "section": "3-000", "section_title": "Maintenance of docket", "comment": "This section speaks of a judge’s duty to maintain the docket. It is actually the court clerk who performs such functions. Recommend this be changed to state the clerk shall perform these functions."},
    {"chapter_title": "Court Administration", "section": "3-204", "section_title": "Disturbance of proceedings", "comment": "Recommend replacing Disturbance of proceedings (3-204) with MTAS model: Contempt of court (3-204).  Contempt of court is punishable by a fine of fifty dollars ($50.00), or such lesser amount as may be imposed in the judge's discretion."},
    {"chapter_title": "Summonses and Subpoenas", "section": "3-000", "section_title": "Failure to Appear", "comment": "This section speaks of the power to issue a warrant for arrest. City court judges do not have the power to issue arrest warrants, and no person may be arrested for violation of a municipal ordinance. Recommend language about arrest warrants be deleted from this section."},
    {"chapter_title": "Summonses and Subpoenas", "section": "3-000", "section_title": "Issuance of arrest warrants; issuance of summonses; issuance of subpoenas.", "comment": "Recommend attorney review of these sections as persons may not be arrested for ordinance violations."},
    {"chapter_title": "Warrants", "section": "3-300", "section_title": None, "comment": "Recommend removing “Warrants” from chapter title for reasons provided in § 3-301 below."},
    {"chapter_title": None, "section": "2-000", "section_title": None, "comment": "If current code is pre-Municipal Court Reform Act 2004, Recommend adopting the MTAS model code title."},
    {"chapter_title": None, "section": "3-102", "section_title": "Jurisdiction", "comment": "Recommend adding this section from MTAS model to Title 3 Chapter 1: Jurisdiction (3-102).  The city judge shall have the authority to try persons charged with the violation of municipal ordinances, and to punish persons convicted of such violations by levying a civil penalty under the general penalty provision of this code."},
    {"chapter_title": None, "section": "3-201", "section_title": None, "comment": "Recommend deleting \"whether committed to workhouse\" as one cannot be imprisoned for an ordinance violation."},
    {"chapter_title": None, "section": "3-202", "section_title": "E-citation fee", "comment": "Recommend deleting language related to e-citation fee as that provision has met the five-year sunset."},
    {"chapter_title": None, "section": "3-202", "section_title": None, "comment": "Recommend modifying to state: “All penalties and costs shall be imposed by the city judge and recorded by the court clerk on the city court docket in open court.” This reflects the court clerk recording on the docket, not the judge."},
    {"chapter_title": None, "section": "3-204", "section_title": None, "comment": "Recommend current “§ 3-204, Disturbance of proceedings” be replaced with model section- “Contempt of court. Contempt of court is punishable by a fine of fifty dollars ($50.00), or such lesser amount as may be imposed in the judge's discretion.”"},
    {"chapter_title": None, "section": "3-205", "section_title": "Trial and disposition of cases", "comment": "Recommend deleting Trial and disposition of cases (3-205). City should remove this section as it is not the actual practice. City schedules cases for trial."},
    {"chapter_title": None, "section": "3-206", "section_title": "E-citations", "comment": "Recommend including the sunset period- 5 years- using the model footnote. E-citations automatically sunset after 5 years and cannot be re-established by ordinance."},
    {"chapter_title": None, "section": "3-301", "section_title": "Issue of arrest warrants", "comment": "Recommend deleting section “Issue of arrest warrants” as you can’t be arrested for municipal ordinance violations."}
  ],
  "4": [
    {"chapter_title": "Infectious Disease Control", "section": "4-000", "section_title": None, "comment": "Infectious disease control policy. Recommend deleting this chapter as it appears obsolete and may be superseded by state and federal laws and regulations."},
    {"chapter_title": "Occupational Safety and Health Program", "section": "4-000", "section_title": None, "comment": "Recommend the MTAS model chapter on OSHA where text references state law."},
    {"chapter_title": "Occupational Safety and Health Program", "section": "4-200", "section_title": "entire chapter", "comment": "Recommend replacing \"Occupational Safety and Health Program\" with MTAS model that follows OSHA by reference."},
    {"chapter_title": "Personnel Regulation", "section": "4-100", "section_title": "entire chapter", "comment": "Recommend using language from model code chapter \"Personnel Regulation\"."},
    {"chapter_title": "Personnel Regulations", "section": "4-000", "section_title": None, "comment": "Recommend maintaining personnel rules and regulations in the office of the recorder and include by reference in the code."},
    {"chapter_title": "Travel Reimbursement Regulations", "section": "4-000", "section_title": "entire chapter", "comment": "Recommend adoption of MTAS model “Travel Reimbursement Regulations” chapter to be compliant with state law. Please choose federal or state rates."},
    {"chapter_title": "Travel Reimbursements", "section": "4-000", "section_title": None, "comment": "Has the city adopted travel reimbursement regulations pursuant to TCA § 6-54-901 – 907?"},
    {"chapter_title": None, "section": "4-000", "section_title": "Social Security", "comment": "This title often provides for the execution of social security agreements with the federal government, but it is no longer necessary for municipalities to go through this process. Recommend deleting this chapter as federal law no longer prescribes for the municipal adoption of the social security system. However, the ordinance(s) involved generally does no harm."},
    {"chapter_title": None, "section": "4-304", "section_title": "Travel", "comment": "Would you prefer to use the federal or state travel reimbursement rates?"},
    {"chapter_title": None, "section": "4-309", "section_title": "Outside Employment", "comment": "Recommend replacing current \"§ 4-309, Outside employment\" with model section."}
  ],
  "5": [
    {"chapter_title": "Privilege Taxes", "section": "5-000", "section_title": "entire chapter", "comment": "LR: Recommend replacing Privilege Taxes (Ch. 3) with the MTAS model for updating cite and minor verbiage changes."},
    {"chapter_title": "Sales Tax", "section": "5-000", "section_title": None, "comment": "Sales Tax. Please check and make certain the amounts stated have not changed since 1996, the last time this chapter was updated."},
    {"chapter_title": "Unclaimed Property", "section": "5-600", "section_title": "entire chapter", "comment": "Recommend deleting chapter on \"Unclaimed Property\" as this is governed by state law and statutes cited herein have been widely amended since the ordinance was adopted."},
    {"chapter_title": None, "section": "5-101", "section_title": None, "comment": "Please provide the bank and city where the bank/branch is located that serves as the official depository for town funds. There may be more than one financial institution listed."},
    {"chapter_title": None, "section": "5-201", "section_title": None, "comment": "When are taxes due and payable? (model provides first Monday of October of the year for which levied) Where are they paid? (draft code provides at the Office of the X County Trustee)"},
    {"chapter_title": None, "section": "5-203", "section_title": "When delinquent - interest charged", "comment": "LR: When delinquent-interest charged (5-203). The footnote needs to be replaced to reflect change in the statute. It should read: Tennessee Code Annotated, section 67-5-2010(b) provides that if the county trustee collects the municipality's property taxes, interest of one and one-half percent (1.5%) shall be added on the first day of March, following the tax due date and on the first day of each succeeding month."},
    {"chapter_title": None, "section": "5-203", "section_title": "When due and payable", "comment": "Recommend adding the following sentence to section “when due and payable”: Taxes may be paid at the Office of the ______________ County Trustee (or city/town clerk)."},
    {"chapter_title": None, "section": "5-301", "section_title": "Tax levied", "comment": "Recommend replacing \"Tax levied\" with MTAS model which references state law instead of reciting the percentages which may be changed periodically by the legislature."}
  ],
  "6": [
    {"chapter_title": "Workhouse", "section": "6-000", "section_title": None, "comment": "Recommend deleting chapter \"Workhouse\" as one cannot be jailed or placed in a workhouse for municipal ordinance violations."},
    {"chapter_title": None, "section": "6-000", "section_title": "entire chapter", "comment": "Do you want to use the police department chapter or the alternate that designates the sheriff's department? Is there an interlocal agreement with the county?"},
    {"chapter_title": None, "section": "6-000", "section_title": "entire title", "comment": "Recommend replacing Title 6 with MTAS model Title 6, Law Enforcement. The MTAS model is three chapters and current with law changes."},
    {"chapter_title": None, "section": "6-103", "section_title": "Policemen to wear uniforms and be armed", "comment": "Recommend deleting section \"Policemen to wear uniforms and be armed.\" as this is better suited for policy than ordinance."},
    {"chapter_title": None, "section": "6-105", "section_title": "Police officers may require assistance in making arrest.", "comment": "Recommend deleting “6-105. Police officers may require assistance in making arrest.”"},
    {"chapter_title": None, "section": "6-105", "section_title": "Policemen may require assistance", "comment": "Recommend deleting \"Policemen may require assistance.\"  because the serious potential liability problems that it could create."},
    {"chapter_title": None, "section": "6-106", "section_title": "Disposition of persons arrested", "comment": "Recommend replacing “6-106. Disposition of persons arrested” with version from MTAS model 6-202."},
    {"chapter_title": None, "section": "6-106", "section_title": None, "comment": "Recommend using model § 6-202 as there are not arrests for municipal ordinance violations. The procedure in place here is not applicable."},
    {"chapter_title": None, "section": "6-107", "section_title": "Police department records", "comment": "Recommend replacing “6-107. Police department records” with MTAS model 6-103."},
    {"chapter_title": None, "section": "6-303", "section_title": "Automatic dialing devices", "comment": "Recommend deleting \"Automatic dialing devices\" section as telecom is generally a matter of state/federal law, not municipal. TCA § 7-86-118 states in part \"The board of directors of an emergency communications district may, by resolution, vote to preclude service users from programming the emergency number “911” in automatic dialers used in conjunction with security alarm systems.\" I do not see anywhere this power is granted to municipalities."}
  ],
  "7": [
    {"chapter_title": None, "section": "7-101", "section_title": "Fire code adopted", "comment": "Recommend replacing “7-101. Fire code adopted” with MTAS model 7-101. (The language re NFPA Life Safety Code can be deleted if that is the city’s preference.)"},
    {"chapter_title": None, "section": "7-101", "section_title": None, "comment": "We selected the most current editions of the codes cited in this section. Please let us know if earlier versions are preferred."},
    {"chapter_title": None, "section": "7-103", "section_title": None, "comment": "Please provide any modifications required."},
    {"chapter_title": None, "section": "7-200", "section_title": None, "comment": "Is there an employed or volunteer fire department? Use the concise or expanded section?"},
    {"chapter_title": None, "section": "7-300", "section_title": "entire chapter", "comment": "Please choose the fire service option for chapter 3."},
    {"chapter_title": "False Alarms", "section": "False alarms", "section_title": None, "comment": "Suggestion deletion as is above a Class C misdemeanor"}
  ],
  "8": [
    {"chapter_title": "Intoxicating Liquors", "section": "8-000", "section_title": "entire chapter", "comment": "LR: MTAS recommends updating the entire chapter using the MTAS model on Intoxicating Liquors. Town specific items such as \"Maximum number of licenses\" and \"Location of liquor store\" can be incorporated into the revised model."},
    {"chapter_title": "Intoxicating Liquors", "section": "8-100", "section_title": None, "comment": "Recommend replacing the current chapter on Intoxicating Liquors with the MTAS model as modified by certain ordinance provisions. This version will not contain fees for applications for certificates necessary to obtain retail liquor licenses from ABC or bonds for licensees as there is no authority for those fees/bonds in the T.C.A. This version also removes provisions that are in the exclusive jurisdiction of the state to police liquor licensure."},
    {"chapter_title": None, "section": "8-102", "section_title": "Application for certificate", "comment": "Recommend replacing “8-102. Application for Certificate” with MTAS model 8-102."},
    {"chapter_title": None, "section": "8-206", "section_title": "Beer", "comment": "Recommend using model definition of beer so that percentage doesn't require updating as state law is changed. \"'Beer' defined. The term 'beer' as used in this chapter shall be the same definition appearing in Tennessee Code Annotated, § 57-5-101.\""},
    {"chapter_title": None, "section": "8-210", "section_title": "Number of beer permits", "comment": "Would the town like to limit the number of beer permits or leave it open ended?"}
  ],
  "9": [
    {"chapter_title": "Adult Oriented Businesses", "section": "9-900", "section_title": None, "comment": "Recommend using MTAS model on Adult Oriented Businesses for updated language."},
    {"chapter_title": "Cable Television", "section": "9-301", "section_title": None, "comment": "Do you have a local franchise agreement for cable TV or is it provided under the state franchise agreement? If it is provided by local agreement, what is the ordinance number and date of passage?"},
    {"chapter_title": "Cable Television", "section": "9-600", "section_title": None, "comment": "Recommend using model code \"Cable Television\"."},
    {"chapter_title": "Charitable Solicitors", "section": "9-000", "section_title": None, "comment": "Recommend replacing Title 9 Chapter 2 “Peddlers, ETC” and Title 9 Chapter 3 “Charitable Solicitors” with MTAS model Title 9 Chapter 1 “Peddlers, Solicitors, ETC.”"},
    {"chapter_title": "Mobile Home Parks", "section": "9-000", "section_title": "entire chapter", "comment": "Recommend chapter on Mobile Home Parks be moved to Title 14 on Zoning and Land Use Control."},
    {"chapter_title": "Peddlers, Etc.", "section": "9-000", "section_title": None, "comment": "Recommend replacing Title 9 Chapter 2 “Peddlers, ETC” and Title 9 Chapter 3 “Charitable Solicitors” with MTAS model Title 9 Chapter 1 “Peddlers, Solicitors, ETC.”"},
    {"chapter_title": "Peddlers, Solicitors, Etc.", "section": "9-100", "section_title": None, "comment": "Recommend using model chapter 1 re \"Peddlers, Solicitors, Etc.\" to update chapter for first amendment compliance."},
    {"chapter_title": "Street shows & medicine shows", "section": "9-000", "section_title": None, "comment": "Recommend deleting chapter on \"Street shows & medicine shows\"."},
    {"chapter_title": "Taxis", "section": "9-500", "section_title": None, "comment": "With the prominence of rideshare, do you want/need to keep this chapter on Taxi regulation?"},
    {"chapter_title": "Wholesale Beer Tax", "section": "entire chapter", "section_title": None, "comment": "Recommend adopting the MTAS model “Wholesale Beer Tax” chapter."},
    {"chapter_title": "Yard Sales", "section": "9-700", "section_title": None, "comment": "Recommend using MTAS model on Yard Sales for updated language."},
    {"chapter_title": None, "section": "9-101", "section_title": "Street barker", "comment": "Recommend deleting \"street barker\" for 1st Amendment concerns."},
    {"chapter_title": None, "section": "9-500", "section_title": "cable franchise", "comment": "Is cable franchise agreement current? If operating under state franchise agreement, Recommend deleting chapter."},
    {"chapter_title": "Pool Rooms", "section": "9-000", "section_title": "entire chapter", "comment": "Recommend chapter on pool halls be deleted or, at the least, section entitled \"Minors to be kept out; exception' be deleted."}
  ],
  "10": [
    {"chapter_title": "Dogs and Cats", "section": "10-200", "section_title": "entire chapter", "comment": "Recommend using the model for updated language for “dogs and cats” chapter."},
    {"chapter_title": "In general", "section": "10-100", "section_title": "entire chapter", "comment": "Recommend using the model for updated language for “in general” chapter."},
    {"chapter_title": None, "section": "10-000", "section_title": "entire title", "comment": "Recommend using the model for updated language for for title which includes “dogs and cats” and \"in general\" chapters."},
    {"chapter_title": None, "section": "10-000", "section_title": None, "comment": "Used MTAS model title to bring up-to-date."}
  ],
  "11": [
    {"chapter_title": "Ephedrine and Ephedrine Related Products", "section": "11-900", "section_title": "entire chapter", "comment": "Recommend deleting chapter on \"Ephedrine and Ephedrine Related Products\" as they are outside of municipal jurisdiction."},
    {"chapter_title": "Firearms, Weapons, and Missiles", "section": "11-600", "section_title": "entire chapter", "comment": "Recommend deleting chapter on \"Firearms, Weapons, and Missiles\" as these would rise above a Class C misdemeanor and/or are pre-empted by state law."},
    {"chapter_title": "Fortune telling, etc.", "section": "11-200", "section_title": "entire chapter", "comment": "Recommend deleting chapter \"fortune telling, etc.\" due to free speech concerns."},
    {"chapter_title": "Handbills", "section": "11-000", "section_title": "entire chapter", "comment": "Recommend deleting chapter containing the Handbill Ordinance as there is TN Supreme Court case law on handbills and this ordinance does not appear to overcome that challenge. Also, if littering is argued, that is above a Class C misdemeanor."},
    {"chapter_title": "Interference with Public Operations and Personnel", "section": "11-400", "section_title": "entire chapter", "comment": "Recommend deleting chapter on \"Interference with Public Operations and Personnel\" as these would rise above a Class C misdemeanor."},
    {"chapter_title": "Miscellaneous", "section": "11-800", "section_title": "entire chapter", "comment": "Recommend deleting Chapter 8 \"Miscellaneous\" as the items covered are either no longer prosecuted or criminal in nature above a Class C misdemeanor."},
    {"chapter_title": "Offenses against the peace and quiet", "section": "11-400", "section_title": "entire chapter", "comment": "Recommend using the model \"offenses against the peace and quiet\" as current chapter has several areas that may violate freedom of speech."},
    {"chapter_title": "Offenses against the person", "section": "11-300", "section_title": "entire chapter", "comment": "Recommend deleting chapter \"offenses against the person\" as assault and battery are above class C misdemeanors."},
    {"chapter_title": "Trespassing and Interference with Traffic", "section": "11-700", "section_title": "entire chapter", "comment": "Recommend using model chapter on \"Trespassing and Interference with Traffic\" as it updates the language and removes malicious mischief which is above a Class C misdemeanor."},
    {"chapter_title": None, "section": "11-102", "section_title": "Minors in beer places", "comment": "Recommend deleting “11-102. Minors in beer places” as it is covered in Title 8."},
    {"chapter_title": None, "section": "11-201", "section_title": "Disturbing the peace", "comment": "Recommend deleting “11-201. Disturbing the Peace.\" This appears criminal in nature, and each crime prohibited is classified as being above a Class C misdemeanor in state law."},
    {"chapter_title": None, "section": "11-202", "section_title": "Disturbing public worship", "comment": "Recommend deleting “11-202. Disturbing public worship\" section as it could violate 1st Amendment protections."},
    {"chapter_title": None, "section": "11-502", "section_title": "Trespassing on trains", "comment": "Recommend deleting “11-502. Trespassing on trains\" as trains are regulated under federal law."},
    {"chapter_title": None, "section": "11-601", "section_title": "Abandoned refrigerators, etc.", "comment": "Recommend deleting ‘11-601. Abandoned refrigerators, etc.\" This section prohibits leaving a refrigerator abandoned or unattended without first removing the door or latch. This is similar to the criminal offense of leaving airtight containers or a refrigerator outside a dwelling without removing the door or latch which, under Tennessee Code Annotated, § 39-17-103, is a class B misdemeanor. Municipal courts do not have jurisdiction to try class B misdemeanors."},
    {"chapter_title": None, "section": "11-602", "section_title": "Caves, wells, cisterns, etc.", "comment": "Recommend deleting “11-602. Caves, wells, cisterns, etc.\" This section prohibits activity rarely punished and is out of date."},
    {"chapter_title": None, "section": "11-603", "section_title": "Posting notices, etc.", "comment": "Recommend deleting “11-603. Posting notices, etc.\" This constitutes vandalism, which is a class A misdemeanor or class E felony, per Tennessee Code Annotated, § 39-14-408. The city may not make any state crime an ordinance violation if the state crime is anything other than a class C misdemeanor."},
    {"chapter_title": "Municipal Offenses", "section": "11-", "section_title": "Offenses against administration of government", "comment": "Chapter deleted as these would rise above a Class C misdemeanor"},
    {"chapter_title": "Municipal Offenses", "section": "11-", "section_title": "Offenses against property", "comment": "Chapter deleted as these would rise above a Class C misdemeanor."},
    {"chapter_title": "Municipal Offenses", "section": "11-", "section_title": "Offenses against public health", "comment": "Chapter deleted as offenses are a state crime or rarely punished."}
  ],
  "12": [
    {"chapter_title": "Administrative Hearing Officer", "section": "12-000", "section_title": None, "comment": "Recommend replacing Title 12 Chapter 10 \"Administrative Hearing Officer\" with MTAS model Title 12 Chapter 11 \"Administrative Hearing Officer\" as it is more complete."},
    {"chapter_title": "Public nuisances", "section": "12-000", "section_title": "entire chapter", "comment": "Recommend deleting chapter on public nuisances as I don't see authority for it in the TCA as written and overgrown lots are covered in Title 13."},
    {"chapter_title": None, "section": "12-000", "section_title": "entire chapter", "comment": "For either Title 12 option we will need the year version you would like to use. We recommend the newest edition shown."},
    {"chapter_title": None, "section": "12-000", "section_title": None, "comment": "Recommend using the short option for Title 12 at the end of the draft that incorporates the codes by reference instead of the multi-chapter version."}
  ],
  "13": [
    {"chapter_title": "Junkyards", "section": "12-000", "section_title": "entire chapter", "comment": "Chapter 3 on Junkyards has two options. Please select the appropriate option for you."},
    {"chapter_title": "Junkyards", "section": "13-", "section_title": "Junk and debris on private property prohibited.", "comment": "Delete any language re: city removing at the expense of owner/occupant of property as city has no authority to enter property in this manner. Additional remedy available in T. 13 Ch. 4 (junked motor vehicles)."},
    {"chapter_title": "Slum Clearance", "section": "13-000", "section_title": "entire chapter", "comment": "Recommend adding MTAS model chapter regarding \"Slum Clearance\"."},
    {"chapter_title": None, "section": "13-305", "section_title": "Penalty", "comment": "Recommend replacing “Penalty” paragraph with MTAS model to bring it in compliance with state law regarding fines:  Any person violating this chapter shall be subject to a civil penalty of fifty dollars ($50.00) plus court costs for each separate violation of this chapter.  In addition, pursuant to Tennessee Code Annotated, § 55-5-122, the municipal court may issue an order to remove vehicles from private property. Each day the violation of this chapter continues shall be considered a separate violation."}
  ],
  "14": [
    {"chapter_title": "Trailer Parks", "section": "14-000", "section_title": "entire chapter", "comment": "Recommend replacing chapter 5 “Trailer Parks” with MTAS model chapter “Mobile Home Parks”."},
    {"chapter_title": "Zoning", "section": "14-200", "section_title": None, "comment": "Zoning is included in the code by reference. Please supply the original zoning ordinance number (and provide the ordinance)."},
    {"chapter_title": None, "section": "14-101", "section_title": None, "comment": "Determine payment, if any for the planning commission."},
    {"chapter_title": None, "section": "14-103", "section_title": "Powers and duties", "comment": "Recommend replacing “14-103. Powers and duties\" with MTAS model 14-102 \"Organization, powers, duties, etc.\""},
    {"chapter_title": None, "section": "14-303", "section_title": None, "comment": "Please provide panel numbers for blanks."}
  ],
  "15": [
    {"chapter_title": "Turning Movements", "section": "15-000", "section_title": "entire chapter", "comment": "Recommend replacing Title 15 Chapter 4 \"Turning Movements\" with MTAS model Title 15 Chapter 4 \"Turning Movements.\""},
    {"chapter_title": None, "section": "15-103", "section_title": "Careless driving", "comment": "Recommend deleting section as careless driving is above a Class C misdemeanor and not in municipal court jurisdiction."},
    {"chapter_title": None, "section": "15-103", "section_title": "Reckless driving", "comment": "Recommend deleting section on Reckless driving as it is above a class C misdemeanor and not actionable in a civil municipal court."},
    {"chapter_title": None, "section": "15-109", "section_title": "General requirements for traffic control signs, etc.", "comment": "Recommend replacing “General requirements for traffic control signs, etc.” with MTAS model language and accompanying footnote: Pursuant to Tennessee Code Annotated, § 54-5-108, all traffic control signs, signals, markings, and devices shall conform to the latest revision of the Tennessee Manual on Uniform Traffic Control Devices for Streets and Highways,2 and shall be uniform as to type and location throughout the city/town. FN: 2For the latest revision of the Tennessee Manual on Uniform Traffic Control Devices for Streets and Highways, see the Official Compilation of the Rules and Regulations of the State of Tennessee, § 1680-3-1, et seq."},
    {"chapter_title": None, "section": "15-119", "section_title": "Vehicles and operators to be licensed", "comment": "Recommend replacing Vehicles and operators to be licensed with MTAS model “It shall be unlawful for any person to operate a motor vehicle in violation of the \"Tennessee Motor Vehicle Title and Registration Law\" or the \"Uniform Classified and Commercial Driver License Act of 1988.\"”"},
    {"chapter_title": None, "section": "15-122", "section_title": "Bicycle riders, etc.", "comment": "Recommend replacing section “Bicycle riders, etc.” with MTAS model “Motorcycles, motor driven cycles, motorized bicycles, bicycles, etc."},
    {"chapter_title": None, "section": "15-123", "section_title": "Compliance with financial responsibility law required", "comment": "Recommend replacing \"Compliance with financial responsibility law required\" with MTAS model for most current version."},
    {"chapter_title": None, "section": "15-202", "section_title": "Operation of authorized emergency vehicles", "comment": "Recommend replacing \"Operation of authorized emergency vehicles\" with MTAS model for updated verbiage."},
    {"chapter_title": None, "section": "15-303", "section_title": "In school zones", "comment": "Recommend replacing \"In school zones\" with MTAS model."},
    {"chapter_title": None, "section": "15-503", "section_title": "Upon approach of authorized emergency vehicle", "comment": "Recommend deleting \"Upon approach of authorized emergency vehicle\" as offense is above a class C misdemeanor."},
    {"chapter_title": None, "section": "15-504", "section_title": "At railroad crossings", "comment": "Recommend replacing \"At railroad crossings\" with MTAS model."},
    {"chapter_title": None, "section": "15-506", "section_title": "At yield signs", "comment": "Recommend replacing \"At yield signs\" with MTAS model."},
    {"chapter_title": None, "section": "15-507", "section_title": "At traffic-control signals generally", "comment": "Recommend replacing \"At traffic-control signals generally\" with MTAS model."},
    {"chapter_title": None, "section": "15-510", "section_title": "Stops to be signaled", "comment": "Recommend replacing \"Stops to be signaled\" with MTAS model."},
    {"chapter_title": None, "section": "15-706", "section_title": "Deposit of license in lieu of bail", "comment": "Recommend deleting “15-706. Deposit of license in lieu of bail\" as the referenced statutes were repealed by the legislature."},
    {"chapter_title": "Motor vehicles, traffic & parking", "section": "15-", "section_title": "Presumption with respect to traffic control signals", "comment": "Section deleted as this raises liability for fraudulently placed traffic control signals."}
  ],
  "16": [
    {"chapter_title": None, "section": "16-107", "section_title": "Littering streets, alleys, or sidewalks prohibited", "comment": "Recommend deleting “16-107. Littering streets, alleys, or sidewalks prohibited.\" Littering is a class B misdemeanor, per Tennessee Code Annotated, § 39-14-502. The town may not make any state crime an ordinance violation if the state crime is anything other than a class C misdemeanor."},
    {"chapter_title": None, "section": "16-203", "section_title": "Fee", "comment": "Recommend replacing “16-203. Fee\" with MTAS model §16-203 \"Fee\" as the city's current fee structure is very low."},
    {"chapter_title": None, "section": "16-203", "section_title": "Fee", "comment": "Recommend replacing \"Fee\" with MTAS model and inserting \"Deposit or bond\" from MTAS model immediately after."},
    {"chapter_title": None, "section": "16-207", "section_title": "Insurance", "comment": "Recommend replacing \"Insurance\" with MTAS model."},
    {"chapter_title": None, "section": "16-209", "section_title": "Enforcement and penalties", "comment": "Recommend replacing \"Enforcement and penalties\" with MTAS model \"Violations and penalty\"."}
  ],
  "17": [
    {"chapter_title": None, "section": "17-000", "section_title": None, "comment": "Do you have a contract for garbage service or do the residents contract individually?"},
    {"chapter_title": None, "section": "17-112", "section_title": "Violations and penalty", "comment": "Recommend adding new section on penalties as last section in chapter. “Violations and penalty. Violations of this chapter shall subject the offender to a penalty under the general penalty provision of this code.  Each day a violation is allowed to continue shall constitute a separate offense.” Also Recommend allowing MTAS to add this language to the end of other chapters as it sees proper."}
  ],
  "18": [
    {"chapter_title": None, "section": "18-000", "section_title": "entire title", "comment": "Recommend replacing Title 18 with MTAS model Title 18."},
    {"chapter_title": "Water & Sewers", "section": "Fluoridization of water", "section_title": "Fluoridization of water", "comment": "Section deleted as the process is governed under state law"}
  ],
  "19": [
    {"chapter_title": None, "section": "19-000", "section_title": None, "comment": "Please review Title 19 (Electricity and Gas) for outdated utility franchise language and confirm whether any chapters should be updated to reflect current franchise agreements or state law."},
    {"chapter_title": None, "section": "19-000", "section_title": None, "comment": "Confirm whether the city provides electricity/gas service directly or through a utility provider, as this affects which model code chapters apply."},
    {"chapter_title": None, "section": "19-000", "section_title": None, "comment": "Review any referenced rate schedules or tariff provisions to confirm they are current."}
  ],
  "20": [
    {"chapter_title": "Fair Housing Regulations", "section": "20-000", "section_title": "entire chapter", "comment": "Recommend deleting chapter 1 \"Fair Housing Regulations\" as this is enforceable on the federal level."},
    {"chapter_title": "Fair Housing Regulations", "section": "20-000", "section_title": None, "comment": "Recommend deleting Title 20 Chapter 2 \"Fair Housing Regulations\" as this is covered in federal law and penalty described is not allowable for cities."}
  ]
}

TITLE_NAMES = {
    1: "General Administration",
    2: "Boards and Commissions, etc.",
    3: "Municipal Court",
    4: "Municipal Personnel",
    5: "Municipal Finance and Taxation",
    6: "Law Enforcement",
    7: "Fire Protection and Fireworks",
    8: "Alcoholic Beverages",
    9: "Business, Peddlers, Solicitors, etc.",
    10: "Animal Control",
    11: "Municipal Offenses",
    12: "Building, Utility, etc. Codes",
    13: "Property Maintenance Regulations",
    14: "Zoning and Land Use Control",
    15: "Motor Vehicles, Traffic and Parking",
    16: "Streets and Sidewalks, etc.",
    17: "Refuse and Trash Disposal",
    18: "Water and Sewers",
    19: "Electricity and Gas",
    20: "Miscellaneous",
}

# ---------------------------------------------------------------------------
# Claude API
# ---------------------------------------------------------------------------

def build_prompt(title_num: int, city_name: str, city_code: str, mtas_text) -> str:
    title_name = TITLE_NAMES.get(title_num, f"Title {title_num}")
    issues = CHECKLIST.get(str(title_num), [])

    checklist_lines = []
    for i, item in enumerate(issues, 1):
        parts = [f"{i}."]
        if item.get("chapter_title"):
            parts.append(f"Chapter: {item['chapter_title']}")
        if item.get("section"):
            parts.append(f"Section: {item['section']}")
        if item.get("section_title"):
            parts.append(f"Title: {item['section_title']}")
        parts.append(f"Standard comment: {item['comment']}")
        checklist_lines.append("  " + " | ".join(parts))

    checklist_block = "\n".join(checklist_lines) if checklist_lines else "  (no standard issues for this title)"

    mtas_block = ""
    if mtas_text:
        mtas_block = f"""
MTAS MODEL CODE — TITLE {title_num}:
{mtas_text.strip()}
"""

    return f"""You are performing a legal review of Title {title_num} ({title_name}) of {city_name}'s municipal code against MTAS (Municipal Technical Advisory Service) standards for Tennessee municipalities.

STANDARD ISSUES CHECKLIST FOR TITLE {title_num}:
{checklist_block}
{mtas_block}
CITY CODE — {city_name.upper()} — TITLE {title_num}:
{city_code.strip()}

INSTRUCTIONS:

STEP 1 — FILTER THE CHECKLIST (critical):
Read each numbered checklist item carefully. Ask yourself: does the city's actual code contain the specific problematic section, language, or pattern described?
  - INCLUDE the item ONLY if the exact problem is present in the city's code.
  - DO NOT include an item if: the city does not have that section at all, the city already uses the recommended language, or the issue simply does not arise in this city's code.
  - DO NOT include an item just because the topic exists; the specific defect described must be present.

STEP 2 — EXTRACT SECTION TEXT:
For each item you include, copy the relevant verbatim text from the city's code (up to ~300 words). If the section is absent, set section_text to null.

STEP 3 — NEW FINDINGS:
Identify issues in the city's code that are NOT covered by the checklist. Use your knowledge of Tennessee municipal law. These are optional — only include genuine problems.

Return a JSON array. Each element must have these exact fields:
  - "title": string (e.g. "3")
  - "chapter_title": string or null
  - "section": string (section number)
  - "section_title": string or null
  - "section_text": string (actual text from city's code, or null if section is absent)
  - "comment": string (use the standard comment verbatim for checklist items; write your own for new findings)
  - "source": "checklist" or "new_finding"

Return ONLY the JSON array, no other text. If no issues apply, return an empty array: []"""


def call_claude(client: anthropic.Anthropic, prompt: str) -> list[dict]:
    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=16000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = response.content[0].text.strip()
    if response.stop_reason == "max_tokens":
        raise ValueError("Response was cut off (max_tokens reached). Title may be too large to process in one call.")
    # strip markdown code fences if present
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        raw = raw.rsplit("```", 1)[0].strip()
    # extract JSON array even if the model added text before/after
    start = raw.find("[")
    end = raw.rfind("]")
    if start == -1 or end == -1:
        raise ValueError(f"No JSON array found in response. Raw response: {raw[:300]}")
    raw = raw[start:end + 1]
    return json.loads(raw)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEW_FINDING_FILL = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = [8, 22, 12, 28, 45, 60, 16]
HEADERS = ["Title", "Chapter Title", "Section", "Section Title", "Section Text", "Comment", "Source"]


def write_excel(findings: list[dict], city_name: str, output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    # header row
    ws.append(HEADERS)
    for col_idx, (cell, width) in enumerate(zip(ws[1], COL_WIDTHS), 1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    for row_idx, f in enumerate(findings, 2):
        row = [
            f.get("title", ""),
            f.get("chapter_title") or "",
            f.get("section") or "",
            f.get("section_title") or "",
            f.get("section_text") or "",
            f.get("comment") or "",
            f.get("source") or "",
        ]
        ws.append(row)
        is_new = f.get("source") == "new_finding"
        fill = NEW_FINDING_FILL if is_new else (ALT_FILL if row_idx % 2 == 0 else None)
        for col_idx, cell in enumerate(ws[row_idx], 1):
            if fill:
                cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["City", city_name])
    ws2.append(["Date", datetime.now().strftime("%Y-%m-%d")])
    ws2.append(["Total findings", len(findings)])
    ws2.append(["Standard checklist hits", sum(1 for f in findings if f.get("source") == "checklist")])
    ws2.append(["New findings", sum(1 for f in findings if f.get("source") == "new_finding")])
    ws2.append([])
    ws2.append(["Title", "Findings"])
    title_counts: dict[str, int] = {}
    for f in findings:
        t = str(f.get("title", ""))
        title_counts[t] = title_counts.get(t, 0) + 1
    for t in sorted(title_counts, key=lambda x: int(x) if x.isdigit() else 99):
        label = f"Title {t} — {TITLE_NAMES.get(int(t), '')}" if t.isdigit() else t
        ws2.append([label, title_counts[t]])
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def load_text(path: Path):
    if path.exists():
        return path.read_text(encoding="utf-8", errors="replace")
    return None


def main():
    parser = argparse.ArgumentParser(description="Municipal Code Review Tool")
    parser.add_argument("--city", required=True, help="City name (e.g. Memphis)")
    parser.add_argument("--code-dir", default=None,
                        help="Directory with city code files named title_N.txt (default: cities/<city>/)")
    parser.add_argument("--mtas-dir", default="mtas",
                        help="Directory with MTAS model files named title_N.txt (default: mtas/)")
    parser.add_argument("--titles", default=None,
                        help="Comma-separated title numbers to review (default: all found)")
    parser.add_argument("--output", default=None,
                        help="Output Excel path (default: output/<city>_review_<date>.xlsx)")
    args = parser.parse_args()

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("Error: ANTHROPIC_API_KEY environment variable not set.")

    base_dir = Path(__file__).parent
    code_dir = Path(args.code_dir) if args.code_dir else base_dir / "cities" / args.city.lower().replace(" ", "_")
    mtas_dir = base_dir / args.mtas_dir
    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)

    if args.output:
        output_path = Path(args.output)
    else:
        date_str = datetime.now().strftime("%Y%m%d")
        output_path = output_dir / f"{args.city.replace(' ', '_')}_review_{date_str}.xlsx"

    if args.titles:
        titles = [int(t.strip()) for t in args.titles.split(",")]
    else:
        # discover all title_N.txt files in code_dir
        if not code_dir.exists():
            sys.exit(f"Error: code directory not found: {code_dir}\nCreate it and add title_N.txt files, or use --code-dir.")
        titles = sorted(
            int(f.stem.split("_")[1])
            for f in code_dir.glob("title_*.txt")
            if f.stem.split("_")[1].isdigit()
        )
        if not titles:
            sys.exit(f"No title_N.txt files found in {code_dir}")

    client = anthropic.Anthropic(api_key=api_key)
    all_findings: list[dict] = []

    for title_num in titles:
        city_file = code_dir / f"title_{title_num}.txt"
        mtas_file = mtas_dir / f"title_{title_num}.txt"

        city_text = load_text(city_file)
        if not city_text:
            print(f"  [skip] Title {title_num}: no city code file at {city_file}")
            continue

        mtas_text = load_text(mtas_file)
        title_name = TITLE_NAMES.get(title_num, f"Title {title_num}")
        mtas_note = "" if mtas_text else " (no MTAS model file)"
        print(f"  Reviewing Title {title_num} — {title_name}{mtas_note}...", end=" ", flush=True)

        prompt = build_prompt(title_num, args.city, city_text, mtas_text)
        try:
            findings = call_claude(client, prompt)
            all_findings.extend(findings)
            print(f"{len(findings)} finding(s)")
        except Exception as e:
            print(f"ERROR: {e}")

    if not all_findings:
        print("No findings produced. Check your input files.")
        return

    write_excel(all_findings, args.city, output_path)
    print(f"\nDone. {len(all_findings)} total findings -> {output_path}")


if __name__ == "__main__":
    main()
